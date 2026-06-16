import os
from datetime import datetime, timedelta
from typing import List, Optional, Dict
from sqlalchemy.orm import Session
from .database import get_db, config, BASE_DIR
from .models import Server, Alert, ExpansionPlan, HealthReport, ResourceMetric, Verification, PurchaseOrder, Approval
from . import audit_logger


class ReportGenerator:
    def __init__(self, db: Session = None):
        self.db = db or next(get_db())
        self.export_dir = os.path.join(BASE_DIR, config['report']['export_dir'])
        os.makedirs(self.export_dir, exist_ok=True)

    def generate_daily_report(self, report_date: datetime = None) -> HealthReport:
        if report_date is None:
            report_date = datetime.now()

        report_date = report_date.replace(hour=0, minute=0, second=0, microsecond=0)

        existing = self.db.query(HealthReport).filter(
            HealthReport.report_date == report_date,
            HealthReport.report_type == "daily"
        ).first()
        if existing:
            return existing

        stats = self._calculate_daily_stats(report_date)

        report = HealthReport(
            report_date=report_date,
            report_type="daily",
            total_servers=stats['total_servers'],
            alert_count=stats['alert_count'],
            warning_count=stats['warning_count'],
            critical_count=stats['critical_count'],
            expansion_count=stats['expansion_count'],
            expansion_completed_count=stats['expansion_completed_count'],
            expansion_completion_rate=stats['expansion_completion_rate'],
            avg_cpu_usage=stats['avg_cpu'],
            avg_memory_usage=stats['avg_memory'],
            avg_disk_usage=stats['avg_disk'],
            avg_network_usage=stats['avg_network'],
            summary=self._generate_summary(stats, report_date)
        )

        self.db.add(report)
        self.db.commit()
        self.db.refresh(report)

        pdf_path = self._export_pdf(report)
        excel_path = self._export_excel(report)

        report.pdf_path = pdf_path
        report.excel_path = excel_path
        self.db.commit()
        self.db.refresh(report)

        audit_logger.log_audit(
            self.db,
            module="report",
            action="generate_daily",
            resource_type="health_report",
            resource_id=report.id,
            operator="system",
            details=f"已生成每日容量健康报告: {report_date.strftime('%Y-%m-%d')}"
        )

        return report

    def _calculate_daily_stats(self, report_date: datetime) -> Dict:
        start_time = report_date
        end_time = report_date + timedelta(days=1)

        total_servers = self.db.query(Server).filter(Server.is_active == True).count()

        alerts = self.db.query(Alert).filter(
            Alert.created_at >= start_time,
            Alert.created_at < end_time
        ).all()

        alert_count = len(alerts)
        warning_count = sum(1 for a in alerts if a.alert_level == 'warning')
        critical_count = sum(1 for a in alerts if a.alert_level in ['critical', 'fatal'])

        expansions = self.db.query(ExpansionPlan).filter(
            ExpansionPlan.created_at >= start_time,
            ExpansionPlan.created_at < end_time
        ).all()

        expansion_count = len(expansions)
        expansion_completed_count = sum(1 for e in expansions if e.status == 'approved')
        expansion_completion_rate = (
            round(expansion_completed_count / expansion_count * 100, 2)
            if expansion_count > 0 else 0.0
        )

        metrics = self.db.query(ResourceMetric).filter(
            ResourceMetric.timestamp >= start_time,
            ResourceMetric.timestamp < end_time
        ).all()

        if metrics:
            avg_cpu = round(sum(m.cpu_usage for m in metrics) / len(metrics), 2)
            avg_memory = round(sum(m.memory_usage for m in metrics) / len(metrics), 2)
            avg_disk = round(sum(m.disk_usage for m in metrics) / len(metrics), 2)
            avg_network = round(sum(m.network_usage for m in metrics) / len(metrics), 2)
        else:
            avg_cpu = avg_memory = avg_disk = avg_network = 0.0

        return {
            'total_servers': total_servers,
            'alert_count': alert_count,
            'warning_count': warning_count,
            'critical_count': critical_count,
            'expansion_count': expansion_count,
            'expansion_completed_count': expansion_completed_count,
            'expansion_completion_rate': expansion_completion_rate,
            'avg_cpu': avg_cpu,
            'avg_memory': avg_memory,
            'avg_disk': avg_disk,
            'avg_network': avg_network
        }

    def _generate_summary(self, stats: Dict, report_date: datetime) -> str:
        health_status = "良好"
        if stats['critical_count'] > 0:
            health_status = "需关注"
        elif stats['warning_count'] > 3:
            health_status = "一般"

        return (
            f"容量健康日报 - {report_date.strftime('%Y年%m月%d日')}\n"
            f"{'='*40}\n"
            f"整体健康状态: {health_status}\n"
            f"监控服务器总数: {stats['total_servers']} 台\n"
            f"今日预警总数: {stats['alert_count']} 条\n"
            f"  - 警告级别: {stats['warning_count']} 条\n"
            f"  - 严重/致命级别: {stats['critical_count']} 条\n\n"
            f"扩容申请数: {stats['expansion_count']} 个\n"
            f"扩容完成数: {stats['expansion_completed_count']} 个\n"
            f"扩容完成率: {stats['expansion_completion_rate']}%\n\n"
            f"平均资源使用率:\n"
            f"  - CPU: {stats['avg_cpu']}%\n"
            f"  - 内存: {stats['avg_memory']}%\n"
            f"  - 磁盘: {stats['avg_disk']}%\n"
            f"  - 网络: {stats['avg_network']}%"
        )

    def _export_pdf(self, report: HealthReport) -> str:
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import cm
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
            from reportlab.lib import colors
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont

            pdf_filename = f"health_report_{report.report_date.strftime('%Y%m%d')}.pdf"
            pdf_path = os.path.join(self.export_dir, pdf_filename)

            doc = SimpleDocTemplate(pdf_path, pagesize=A4,
                                    rightMargin=2*cm, leftMargin=2*cm,
                                    topMargin=2*cm, bottomMargin=2*cm)

            styles = getSampleStyleSheet()

            try:
                pdfmetrics.registerFont(TTFont('SimHei', 'C:/Windows/Fonts/simhei.ttf'))
                title_font = 'SimHei'
                body_font = 'SimHei'
            except:
                title_font = 'Helvetica'
                body_font = 'Helvetica'

            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontName=title_font,
                fontSize=18,
                spaceAfter=20,
                alignment=1
            )

            heading_style = ParagraphStyle(
                'CustomHeading',
                parent=styles['Heading2'],
                fontName=title_font,
                fontSize=14,
                spaceAfter=10,
                spaceBefore=15
            )

            normal_style = ParagraphStyle(
                'CustomNormal',
                parent=styles['Normal'],
                fontName=body_font,
                fontSize=10,
                spaceAfter=5,
                leading=15
            )

            story = []

            story.append(Paragraph("IT容量管理系统 - 健康日报", title_style))
            story.append(Paragraph(f"报告日期: {report.report_date.strftime('%Y年%m月%d日')}", normal_style))
            story.append(Spacer(1, 0.5*cm))

            story.append(Paragraph("一、总体概览", heading_style))

            overview_data = [
                ['指标', '数值'],
                ['监控服务器总数', f"{report.total_servers} 台"],
                ['今日预警总数', f"{report.alert_count} 条"],
                ['警告级别预警', f"{report.warning_count} 条"],
                ['严重/致命级别预警', f"{report.critical_count} 条"],
                ['扩容申请数', f"{report.expansion_count} 个"],
                ['扩容完成数', f"{report.expansion_completed_count} 个"],
                ['扩容完成率', f"{report.expansion_completion_rate}%"],
            ]

            overview_table = Table(overview_data, colWidths=[8*cm, 6*cm])
            overview_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), body_font),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))
            story.append(overview_table)

            story.append(Paragraph("二、资源使用率", heading_style))

            usage_data = [
                ['资源类型', '平均使用率', '状态'],
                ['CPU', f"{report.avg_cpu_usage}%", self._get_usage_status(report.avg_cpu_usage)],
                ['内存', f"{report.avg_memory_usage}%", self._get_usage_status(report.avg_memory_usage)],
                ['磁盘', f"{report.avg_disk_usage}%", self._get_usage_status(report.avg_disk_usage)],
                ['网络带宽', f"{report.avg_network_usage}%", self._get_usage_status(report.avg_network_usage)],
            ]

            usage_table = Table(usage_data, colWidths=[4*cm, 4*cm, 6*cm])
            usage_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), body_font),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))
            story.append(usage_table)

            story.append(Spacer(1, 1*cm))
            story.append(Paragraph("三、报告摘要", heading_style))
            story.append(Paragraph(report.summary.replace('\n', '<br/>'), normal_style))

            story.append(Spacer(1, 1*cm))
            story.append(Paragraph(f"报告生成时间: {report.created_at.strftime('%Y-%m-%d %H:%M:%S')}", normal_style))

            doc.build(story)
            return pdf_path

        except Exception as e:
            print(f"PDF生成失败: {e}")
            return ""

    def _export_excel(self, report: HealthReport) -> str:
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            excel_filename = f"health_report_{report.report_date.strftime('%Y%m%d')}.xlsx"
            excel_path = os.path.join(self.export_dir, excel_filename)

            wb = openpyxl.Workbook()

            ws1 = wb.active
            ws1.title = "总览"

            header_font = Font(bold=True, size=12)
            header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            center_align = Alignment(horizontal='center', vertical='center')
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            ws1['A1'] = "IT容量管理系统 - 健康日报"
            ws1.merge_cells('A1:B1')
            ws1['A1'].font = Font(bold=True, size=14)
            ws1['A1'].alignment = center_align

            ws1['A2'] = f"报告日期: {report.report_date.strftime('%Y年%m月%d日')}"
            ws1.merge_cells('A2:B2')

            overview_data = [
                ['指标', '数值'],
                ['监控服务器总数', f"{report.total_servers} 台"],
                ['今日预警总数', f"{report.alert_count} 条"],
                ['警告级别预警', f"{report.warning_count} 条"],
                ['严重/致命级别预警', f"{report.critical_count} 条"],
                ['扩容申请数', f"{report.expansion_count} 个"],
                ['扩容完成数', f"{report.expansion_completed_count} 个"],
                ['扩容完成率', f"{report.expansion_completion_rate}%"],
            ]

            for i, row in enumerate(overview_data, start=4):
                for j, value in enumerate(row, start=1):
                    cell = ws1.cell(row=i, column=j, value=value)
                    cell.border = thin_border
                    if i == 4:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = center_align

            ws1.column_dimensions['A'].width = 20
            ws1.column_dimensions['B'].width = 20

            ws2 = wb.create_sheet("资源使用率")

            usage_data = [
                ['资源类型', '平均使用率', '状态'],
                ['CPU', report.avg_cpu_usage, self._get_usage_status(report.avg_cpu_usage)],
                ['内存', report.avg_memory_usage, self._get_usage_status(report.avg_memory_usage)],
                ['磁盘', report.avg_disk_usage, self._get_usage_status(report.avg_disk_usage)],
                ['网络带宽', report.avg_network_usage, self._get_usage_status(report.avg_network_usage)],
            ]

            for i, row in enumerate(usage_data, start=1):
                for j, value in enumerate(row, start=1):
                    cell = ws2.cell(row=i, column=j, value=value)
                    cell.border = thin_border
                    if i == 1:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = center_align

            ws2.column_dimensions['A'].width = 15
            ws2.column_dimensions['B'].width = 15
            ws2.column_dimensions['C'].width = 15

            ws3 = wb.create_sheet("预警明细")

            alert_data = [['预警ID', '服务器', '资源类型', '级别', '标题', '状态', '创建时间']]

            start_time = report.report_date
            end_time = report.report_date + timedelta(days=1)
            alerts = self.db.query(Alert).filter(
                Alert.created_at >= start_time,
                Alert.created_at < end_time
            ).order_by(Alert.created_at.desc()).all()

            for alert in alerts:
                server = self.db.query(Server).filter(Server.id == alert.server_id).first()
                server_name = server.name if server else "未知"
                alert_data.append([
                    alert.id,
                    server_name,
                    alert.resource_type,
                    alert.alert_level,
                    alert.title,
                    alert.status,
                    alert.created_at.strftime('%Y-%m-%d %H:%M:%S')
                ])

            for i, row in enumerate(alert_data, start=1):
                for j, value in enumerate(row, start=1):
                    cell = ws3.cell(row=i, column=j, value=value)
                    cell.border = thin_border
                    if i == 1:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = center_align

            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
                ws3.column_dimensions[col].width = 18

            wb.save(excel_path)
            return excel_path

        except Exception as e:
            print(f"Excel生成失败: {e}")
            return ""

    def _get_usage_status(self, value: float) -> str:
        if value < 60:
            return "正常"
        elif value < 80:
            return "预警"
        else:
            return "危险"

    def generate_weekly_report(self, report_date: datetime = None) -> HealthReport:
        if report_date is None:
            report_date = datetime.now()

        report_date = report_date.replace(hour=0, minute=0, second=0, microsecond=0)
        start_date = report_date - timedelta(days=6)

        existing = self.db.query(HealthReport).filter(
            HealthReport.report_date == report_date,
            HealthReport.report_type == "weekly"
        ).first()
        if existing:
            return existing

        stats = self._calculate_period_stats(start_date, report_date + timedelta(days=1), "weekly")

        report = HealthReport(
            report_date=report_date,
            report_type="weekly",
            total_servers=stats['total_servers'],
            alert_count=stats['alert_count'],
            warning_count=stats['warning_count'],
            critical_count=stats['critical_count'],
            expansion_count=stats['expansion_count'],
            expansion_completed_count=stats['expansion_completed_count'],
            expansion_completion_rate=stats['expansion_completion_rate'],
            avg_cpu_usage=stats['avg_cpu'],
            avg_memory_usage=stats['avg_memory'],
            avg_disk_usage=stats['avg_disk'],
            avg_network_usage=stats['avg_network'],
            summary=self._generate_period_summary(stats, start_date, report_date, "weekly")
        )

        self.db.add(report)
        self.db.commit()
        self.db.refresh(report)

        pdf_path = self._export_period_pdf(report, start_date, report_date, "weekly")
        excel_path = self._export_period_excel(report, start_date, report_date, "weekly")

        report.pdf_path = pdf_path
        report.excel_path = excel_path
        self.db.commit()
        self.db.refresh(report)

        audit_logger.log_audit(
            self.db,
            module="report",
            action="generate_weekly",
            resource_type="health_report",
            resource_id=report.id,
            operator="system",
            details=f"已生成每周容量健康报告: {start_date.strftime('%Y-%m-%d')} ~ {report_date.strftime('%Y-%m-%d')}"
        )

        return report

    def generate_monthly_report(self, report_date: datetime = None) -> HealthReport:
        if report_date is None:
            report_date = datetime.now()

        report_date = report_date.replace(hour=0, minute=0, second=0, microsecond=0)
        start_date = report_date.replace(day=1)

        existing = self.db.query(HealthReport).filter(
            HealthReport.report_date == report_date,
            HealthReport.report_type == "monthly"
        ).first()
        if existing:
            return existing

        stats = self._calculate_period_stats(start_date, report_date + timedelta(days=1), "monthly")

        report = HealthReport(
            report_date=report_date,
            report_type="monthly",
            total_servers=stats['total_servers'],
            alert_count=stats['alert_count'],
            warning_count=stats['warning_count'],
            critical_count=stats['critical_count'],
            expansion_count=stats['expansion_count'],
            expansion_completed_count=stats['expansion_completed_count'],
            expansion_completion_rate=stats['expansion_completion_rate'],
            avg_cpu_usage=stats['avg_cpu'],
            avg_memory_usage=stats['avg_memory'],
            avg_disk_usage=stats['avg_disk'],
            avg_network_usage=stats['avg_network'],
            summary=self._generate_period_summary(stats, start_date, report_date, "monthly")
        )

        self.db.add(report)
        self.db.commit()
        self.db.refresh(report)

        pdf_path = self._export_period_pdf(report, start_date, report_date, "monthly")
        excel_path = self._export_period_excel(report, start_date, report_date, "monthly")

        report.pdf_path = pdf_path
        report.excel_path = excel_path
        self.db.commit()
        self.db.refresh(report)

        audit_logger.log_audit(
            self.db,
            module="report",
            action="generate_monthly",
            resource_type="health_report",
            resource_id=report.id,
            operator="system",
            details=f"已生成每月容量健康报告: {start_date.strftime('%Y-%m-%d')} ~ {report_date.strftime('%Y-%m-%d')}"
        )

        return report

    def _calculate_period_stats(self, start_time: datetime, end_time: datetime, period_type: str) -> Dict:
        total_servers = self.db.query(Server).filter(Server.is_active == True).count()

        alerts = self.db.query(Alert).filter(
            Alert.created_at >= start_time,
            Alert.created_at < end_time
        ).all()

        alert_count = len(alerts)
        warning_count = sum(1 for a in alerts if a.alert_level == 'warning')
        critical_count = sum(1 for a in alerts if a.alert_level in ['critical', 'fatal'])

        alert_trend = {}
        for alert in alerts:
            day_key = alert.created_at.strftime('%Y-%m-%d')
            if day_key not in alert_trend:
                alert_trend[day_key] = {'total': 0, 'warning': 0, 'critical': 0}
            alert_trend[day_key]['total'] += 1
            if alert.alert_level == 'warning':
                alert_trend[day_key]['warning'] += 1
            elif alert.alert_level in ['critical', 'fatal']:
                alert_trend[day_key]['critical'] += 1

        expansions = self.db.query(ExpansionPlan).filter(
            ExpansionPlan.created_at >= start_time,
            ExpansionPlan.created_at < end_time
        ).all()

        expansion_count = len(expansions)
        expansion_completed_count = sum(1 for e in expansions if e.status == 'approved')
        expansion_rejected_count = sum(1 for e in expansions if e.status == 'rejected')
        expansion_pending_count = sum(1 for e in expansions if e.status == 'pending')
        expansion_completion_rate = (
            round(expansion_completed_count / expansion_count * 100, 2)
            if expansion_count > 0 else 0.0
        )

        approvals = self.db.query(Approval).filter(
            Approval.approved_at >= start_time,
            Approval.approved_at < end_time
        ).all()
        approval_count = len(approvals)

        orders = self.db.query(PurchaseOrder).filter(
            PurchaseOrder.created_at >= start_time,
            PurchaseOrder.created_at < end_time
        ).all()
        order_count = len(orders)
        order_delivered_count = sum(1 for o in orders if o.status in ['delivered', 'completed'])
        order_in_progress_count = sum(1 for o in orders if o.status in ['issued', 'in_progress'])
        order_cancelled_count = sum(1 for o in orders if o.status == 'cancelled')

        verifications = self.db.query(Verification).filter(
            Verification.verified_at >= start_time,
            Verification.verified_at < end_time
        ).all()
        verification_count = len(verifications)
        verification_passed_count = sum(1 for v in verifications if v.status == 'passed')
        verification_failed_count = sum(1 for v in verifications if v.status in ['failed', 'rolled_back'])
        verification_rolled_back_count = sum(1 for v in verifications if v.is_rolled_back)

        metrics = self.db.query(ResourceMetric).filter(
            ResourceMetric.timestamp >= start_time,
            ResourceMetric.timestamp < end_time
        ).all()

        if metrics:
            avg_cpu = round(sum(m.cpu_usage for m in metrics) / len(metrics), 2)
            avg_memory = round(sum(m.memory_usage for m in metrics) / len(metrics), 2)
            avg_disk = round(sum(m.disk_usage for m in metrics) / len(metrics), 2)
            avg_network = round(sum(m.network_usage for m in metrics) / len(metrics), 2)
        else:
            avg_cpu = avg_memory = avg_disk = avg_network = 0.0

        return {
            'total_servers': total_servers,
            'alert_count': alert_count,
            'warning_count': warning_count,
            'critical_count': critical_count,
            'alert_trend': alert_trend,
            'expansion_count': expansion_count,
            'expansion_completed_count': expansion_completed_count,
            'expansion_rejected_count': expansion_rejected_count,
            'expansion_pending_count': expansion_pending_count,
            'expansion_completion_rate': expansion_completion_rate,
            'approval_count': approval_count,
            'order_count': order_count,
            'order_delivered_count': order_delivered_count,
            'order_in_progress_count': order_in_progress_count,
            'order_cancelled_count': order_cancelled_count,
            'verification_count': verification_count,
            'verification_passed_count': verification_passed_count,
            'verification_failed_count': verification_failed_count,
            'verification_rolled_back_count': verification_rolled_back_count,
            'avg_cpu': avg_cpu,
            'avg_memory': avg_memory,
            'avg_disk': avg_disk,
            'avg_network': avg_network
        }

    def _generate_period_summary(self, stats: Dict, start_date: datetime, end_date: datetime, period_type: str) -> str:
        period_name = "周报" if period_type == "weekly" else "月报"
        health_status = "良好"
        if stats['critical_count'] > 0:
            health_status = "需关注"
        elif stats['warning_count'] > 10:
            health_status = "一般"

        lines = [
            f"容量健康{period_name} - {start_date.strftime('%Y年%m月%d日')} 至 {end_date.strftime('%Y年%m月%d日')}",
            f"{'=' * 40}",
            f"整体健康状态: {health_status}",
            f"监控服务器总数: {stats['total_servers']} 台",
            f"",
            f"【预警趋势汇总】",
            f"  预警总数: {stats['alert_count']} 条",
            f"    - 警告级别: {stats['warning_count']} 条",
            f"    - 严重/致命级别: {stats['critical_count']} 条",
            f"  按日趋势:"
        ]

        for day in sorted(stats['alert_trend'].keys()):
            t = stats['alert_trend'][day]
            lines.append(f"    {day}: 共{t['total']}条 (警告{t['warning']}, 严重{t['critical']})")

        lines.extend([
            f"",
            f"【扩容审批进度】",
            f"  扩容方案总数: {stats['expansion_count']} 个",
            f"    - 待审批: {stats['expansion_pending_count']} 个",
            f"    - 已通过: {stats['expansion_completed_count']} 个",
            f"    - 已拒绝: {stats['expansion_rejected_count']} 个",
            f"  审批完成率: {stats['expansion_completion_rate']}%",
            f"  审批记录数: {stats['approval_count']} 条",
            f"",
            f"【采购交付状态】",
            f"  采购订单总数: {stats['order_count']} 个",
            f"    - 进行中: {stats['order_in_progress_count']} 个",
            f"    - 已交付/完成: {stats['order_delivered_count']} 个",
            f"    - 已取消: {stats['order_cancelled_count']} 个",
            f"",
            f"【扩容验证结果】",
            f"  验证记录总数: {stats['verification_count']} 次",
            f"    - 验证通过: {stats['verification_passed_count']} 次",
            f"    - 验证失败: {stats['verification_failed_count']} 次",
            f"    - 已自动回滚: {stats['verification_rolled_back_count']} 次",
            f"",
            f"【平均资源使用率】",
            f"  - CPU: {stats['avg_cpu']}%",
            f"  - 内存: {stats['avg_memory']}%",
            f"  - 磁盘: {stats['avg_disk']}%",
            f"  - 网络: {stats['avg_network']}%"
        ])

        return "\n".join(lines)

    def _export_period_pdf(self, report: HealthReport, start_date: datetime, end_date: datetime, period_type: str) -> str:
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import cm
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
            from reportlab.lib import colors
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont

            period_label = "周报" if period_type == "weekly" else "月报"
            pdf_filename = f"health_report_{period_type}_{report.report_date.strftime('%Y%m%d')}.pdf"
            pdf_path = os.path.join(self.export_dir, pdf_filename)

            doc = SimpleDocTemplate(pdf_path, pagesize=A4,
                                    rightMargin=2*cm, leftMargin=2*cm,
                                    topMargin=2*cm, bottomMargin=2*cm)

            styles = getSampleStyleSheet()

            try:
                pdfmetrics.registerFont(TTFont('SimHei', 'C:/Windows/Fonts/simhei.ttf'))
                title_font = 'SimHei'
                body_font = 'SimHei'
            except:
                title_font = 'Helvetica'
                body_font = 'Helvetica'

            title_style = ParagraphStyle(
                'CustomTitle', parent=styles['Heading1'],
                fontName=title_font, fontSize=18, spaceAfter=20, alignment=1
            )
            heading_style = ParagraphStyle(
                'CustomHeading', parent=styles['Heading2'],
                fontName=title_font, fontSize=14, spaceAfter=10, spaceBefore=15
            )
            normal_style = ParagraphStyle(
                'CustomNormal', parent=styles['Normal'],
                fontName=body_font, fontSize=10, spaceAfter=5, leading=15
            )

            stats = self._calculate_period_stats(start_date, end_date + timedelta(days=1), period_type)
            story = []

            story.append(Paragraph(f"IT容量管理系统 - 健康{period_label}", title_style))
            story.append(Paragraph(
                f"报告周期: {start_date.strftime('%Y年%m月%d日')} 至 {end_date.strftime('%Y年%m月%d日')}",
                normal_style
            ))
            story.append(Spacer(1, 0.5*cm))

            story.append(Paragraph("一、总体概览", heading_style))
            overview_data = [
                ['指标', '数值'],
                ['监控服务器总数', f"{report.total_servers} 台"],
                [f'{period_label}预警总数', f"{report.alert_count} 条"],
                ['警告级别预警', f"{report.warning_count} 条"],
                ['严重/致命级别预警', f"{report.critical_count} 条"],
                ['扩容方案总数', f"{report.expansion_count} 个"],
                ['扩容完成数', f"{report.expansion_completed_count} 个"],
                ['扩容完成率', f"{report.expansion_completion_rate}%"],
            ]
            overview_table = Table(overview_data, colWidths=[8*cm, 6*cm])
            overview_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), body_font),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))
            story.append(overview_table)

            story.append(Paragraph("二、扩容审批与采购交付", heading_style))
            approval_data = [
                ['类别', '数量'],
                ['待审批扩容方案', f"{stats['expansion_pending_count']} 个"],
                ['已通过扩容方案', f"{stats['expansion_completed_count']} 个"],
                ['已拒绝扩容方案', f"{stats['expansion_rejected_count']} 个"],
                ['审批记录数', f"{stats['approval_count']} 条"],
                ['采购订单总数', f"{stats['order_count']} 个"],
                ['采购进行中', f"{stats['order_in_progress_count']} 个"],
                ['采购已交付', f"{stats['order_delivered_count']} 个"],
                ['采购已取消', f"{stats['order_cancelled_count']} 个"],
            ]
            approval_table = Table(approval_data, colWidths=[8*cm, 6*cm])
            approval_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), body_font),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))
            story.append(approval_table)

            story.append(Paragraph("三、扩容验证结果", heading_style))
            verify_data = [
                ['类别', '数量'],
                ['验证总次数', f"{stats['verification_count']} 次"],
                ['验证通过', f"{stats['verification_passed_count']} 次"],
                ['验证失败', f"{stats['verification_failed_count']} 次"],
                ['自动回滚次数', f"{stats['verification_rolled_back_count']} 次"],
            ]
            verify_table = Table(verify_data, colWidths=[8*cm, 6*cm])
            verify_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), body_font),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))
            story.append(verify_table)

            story.append(Paragraph("四、资源使用率", heading_style))
            usage_data = [
                ['资源类型', '平均使用率', '状态'],
                ['CPU', f"{report.avg_cpu_usage}%", self._get_usage_status(report.avg_cpu_usage)],
                ['内存', f"{report.avg_memory_usage}%", self._get_usage_status(report.avg_memory_usage)],
                ['磁盘', f"{report.avg_disk_usage}%", self._get_usage_status(report.avg_disk_usage)],
                ['网络带宽', f"{report.avg_network_usage}%", self._get_usage_status(report.avg_network_usage)],
            ]
            usage_table = Table(usage_data, colWidths=[4*cm, 4*cm, 6*cm])
            usage_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), body_font),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))
            story.append(usage_table)

            story.append(Spacer(1, 1*cm))
            story.append(Paragraph("五、报告摘要", heading_style))
            story.append(Paragraph(report.summary.replace('\n', '<br/>'), normal_style))

            story.append(Spacer(1, 1*cm))
            story.append(Paragraph(f"报告生成时间: {report.created_at.strftime('%Y-%m-%d %H:%M:%S')}", normal_style))

            doc.build(story)
            return pdf_path

        except Exception as e:
            print(f"PDF生成失败: {e}")
            return ""

    def _export_period_excel(self, report: HealthReport, start_date: datetime, end_date: datetime, period_type: str) -> str:
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            period_label = "周报" if period_type == "weekly" else "月报"
            excel_filename = f"health_report_{period_type}_{report.report_date.strftime('%Y%m%d')}.xlsx"
            excel_path = os.path.join(self.export_dir, excel_filename)

            wb = openpyxl.Workbook()
            header_font = Font(bold=True, size=12)
            header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            center_align = Alignment(horizontal='center', vertical='center')
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )

            stats = self._calculate_period_stats(start_date, end_date + timedelta(days=1), period_type)

            ws1 = wb.active
            ws1.title = "总览"
            ws1['A1'] = f"IT容量管理系统 - 健康{period_label}"
            ws1.merge_cells('A1:B1')
            ws1['A1'].font = Font(bold=True, size=14)
            ws1['A1'].alignment = center_align
            ws1['A2'] = f"报告周期: {start_date.strftime('%Y年%m月%d日')} 至 {end_date.strftime('%Y年%m月%d日')}"
            ws1.merge_cells('A2:B2')

            overview_data = [
                ['指标', '数值'],
                ['监控服务器总数', f"{report.total_servers} 台"],
                [f'{period_label}预警总数', f"{report.alert_count} 条"],
                ['警告级别预警', f"{report.warning_count} 条"],
                ['严重/致命级别预警', f"{report.critical_count} 条"],
                ['扩容方案总数', f"{report.expansion_count} 个"],
                ['扩容完成数', f"{report.expansion_completed_count} 个"],
                ['扩容完成率', f"{report.expansion_completion_rate}%"],
            ]
            for i, row in enumerate(overview_data, start=4):
                for j, value in enumerate(row, start=1):
                    cell = ws1.cell(row=i, column=j, value=value)
                    cell.border = thin_border
                    if i == 4:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = center_align
            ws1.column_dimensions['A'].width = 22
            ws1.column_dimensions['B'].width = 20

            ws2 = wb.create_sheet("预警趋势")
            trend_data = [['日期', '预警总数', '警告级别', '严重/致命级别']]
            for day in sorted(stats['alert_trend'].keys()):
                t = stats['alert_trend'][day]
                trend_data.append([day, t['total'], t['warning'], t['critical']])
            for i, row in enumerate(trend_data, start=1):
                for j, value in enumerate(row, start=1):
                    cell = ws2.cell(row=i, column=j, value=value)
                    cell.border = thin_border
                    if i == 1:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = center_align
            for col in ['A', 'B', 'C', 'D']:
                ws2.column_dimensions[col].width = 18

            ws3 = wb.create_sheet("扩容审批与采购")
            approval_data = [
                ['类别', '数量'],
                ['待审批扩容方案', stats['expansion_pending_count']],
                ['已通过扩容方案', stats['expansion_completed_count']],
                ['已拒绝扩容方案', stats['expansion_rejected_count']],
                ['审批记录数', stats['approval_count']],
                ['采购订单总数', stats['order_count']],
                ['采购进行中', stats['order_in_progress_count']],
                ['采购已交付', stats['order_delivered_count']],
                ['采购已取消', stats['order_cancelled_count']],
            ]
            for i, row in enumerate(approval_data, start=1):
                for j, value in enumerate(row, start=1):
                    cell = ws3.cell(row=i, column=j, value=value)
                    cell.border = thin_border
                    if i == 1:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = center_align
            ws3.column_dimensions['A'].width = 22
            ws3.column_dimensions['B'].width = 15

            ws4 = wb.create_sheet("验证结果")
            verify_data = [
                ['类别', '数量'],
                ['验证总次数', stats['verification_count']],
                ['验证通过', stats['verification_passed_count']],
                ['验证失败', stats['verification_failed_count']],
                ['自动回滚次数', stats['verification_rolled_back_count']],
            ]
            for i, row in enumerate(verify_data, start=1):
                for j, value in enumerate(row, start=1):
                    cell = ws4.cell(row=i, column=j, value=value)
                    cell.border = thin_border
                    if i == 1:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = center_align
            ws4.column_dimensions['A'].width = 20
            ws4.column_dimensions['B'].width = 15

            ws5 = wb.create_sheet("预警明细")
            alert_data = [['预警ID', '服务器', '资源类型', '级别', '标题', '状态', '创建时间']]
            alerts = self.db.query(Alert).filter(
                Alert.created_at >= start_date,
                Alert.created_at < end_date + timedelta(days=1)
            ).order_by(Alert.created_at.desc()).all()
            for alert in alerts:
                server = self.db.query(Server).filter(Server.id == alert.server_id).first()
                server_name = server.name if server else "未知"
                alert_data.append([
                    alert.id, server_name, alert.resource_type,
                    alert.alert_level, alert.title, alert.status,
                    alert.created_at.strftime('%Y-%m-%d %H:%M:%S')
                ])
            for i, row in enumerate(alert_data, start=1):
                for j, value in enumerate(row, start=1):
                    cell = ws5.cell(row=i, column=j, value=value)
                    cell.border = thin_border
                    if i == 1:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = center_align
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
                ws5.column_dimensions[col].width = 18

            wb.save(excel_path)
            return excel_path

        except Exception as e:
            print(f"Excel生成失败: {e}")
            return ""

    def get_reports(self, report_type: str = None, start_date: datetime = None,
                    end_date: datetime = None, limit: int = None) -> List[HealthReport]:
        query = self.db.query(HealthReport)
        if report_type:
            query = query.filter(HealthReport.report_type == report_type)
        if start_date:
            query = query.filter(HealthReport.report_date >= start_date)
        if end_date:
            query = query.filter(HealthReport.report_date <= end_date)
        query = query.order_by(HealthReport.report_date.desc())
        if limit:
            query = query.limit(limit)
        return query.all()

    def get_report(self, report_id: int) -> Optional[HealthReport]:
        return self.db.query(HealthReport).filter(HealthReport.id == report_id).first()
