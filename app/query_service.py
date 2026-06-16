import os
from datetime import datetime, timedelta
from typing import List, Dict, Optional
from sqlalchemy.orm import Session
from .database import get_db, BASE_DIR
from .models import Server, Alert, ExpansionPlan, PurchaseOrder, ResourceMetric, AuditLog, Verification
from . import audit_logger


class QueryService:
    def __init__(self, db: Session = None):
        self.db = db or next(get_db())

    def query_alerts(self, server_name: str = None, start_time: datetime = None,
                     end_time: datetime = None, level: str = None, status: str = None,
                     limit: int = 100, offset: int = 0, export_all: bool = False) -> Dict:
        query = self.db.query(Alert).join(Server, Alert.server_id == Server.id)

        if server_name:
            query = query.filter(Server.name.like(f"%{server_name}%"))
        if start_time:
            query = query.filter(Alert.created_at >= start_time)
        if end_time:
            query = query.filter(Alert.created_at <= end_time)
        if level:
            query = query.filter(Alert.alert_level == level)
        if status:
            query = query.filter(Alert.status == status)

        total = query.count()
        if export_all:
            alerts = query.order_by(Alert.created_at.desc()).all()
        else:
            alerts = query.order_by(Alert.created_at.desc()).offset(offset).limit(limit).all()

        return {
            'total': total,
            'items': alerts,
            'page': offset // limit + 1,
            'page_size': limit
        }

    def query_expansions(self, server_name: str = None, start_time: datetime = None,
                         end_time: datetime = None, status: str = None,
                         limit: int = 100, offset: int = 0, export_all: bool = False) -> Dict:
        query = self.db.query(ExpansionPlan).join(Server, ExpansionPlan.server_id == Server.id)

        if server_name:
            query = query.filter(Server.name.like(f"%{server_name}%"))
        if start_time:
            query = query.filter(ExpansionPlan.created_at >= start_time)
        if end_time:
            query = query.filter(ExpansionPlan.created_at <= end_time)
        if status:
            query = query.filter(ExpansionPlan.status == status)

        total = query.count()
        if export_all:
            plans = query.order_by(ExpansionPlan.created_at.desc()).all()
        else:
            plans = query.order_by(ExpansionPlan.created_at.desc()).offset(offset).limit(limit).all()

        return {
            'total': total,
            'items': plans,
            'page': offset // limit + 1,
            'page_size': limit
        }

    def query_orders(self, order_no: str = None, supplier: str = None, status: str = None,
                     start_time: datetime = None, end_time: datetime = None,
                     server_name: str = None,
                     limit: int = 100, offset: int = 0, export_all: bool = False) -> Dict:
        query = self.db.query(PurchaseOrder).join(
            ExpansionPlan, PurchaseOrder.expansion_plan_id == ExpansionPlan.id
        ).join(Server, ExpansionPlan.server_id == Server.id)

        if order_no:
            query = query.filter(PurchaseOrder.order_no.like(f"%{order_no}%"))
        if supplier:
            query = query.filter(PurchaseOrder.supplier.like(f"%{supplier}%"))
        if server_name:
            query = query.filter(Server.name.like(f"%{server_name}%"))
        if status:
            query = query.filter(PurchaseOrder.status == status)
        if start_time:
            query = query.filter(PurchaseOrder.created_at >= start_time)
        if end_time:
            query = query.filter(PurchaseOrder.created_at <= end_time)

        total = query.count()
        if export_all:
            orders = query.order_by(PurchaseOrder.created_at.desc()).all()
        else:
            orders = query.order_by(PurchaseOrder.created_at.desc()).offset(offset).limit(limit).all()

        return {
            'total': total,
            'items': orders,
            'page': offset // limit + 1,
            'page_size': limit
        }

    def query_verifications(self, server_name: str = None, start_time: datetime = None,
                             end_time: datetime = None, status: str = None,
                             limit: int = 100, offset: int = 0, export_all: bool = False) -> Dict:
        query = self.db.query(Verification).join(
            ExpansionPlan, Verification.expansion_plan_id == ExpansionPlan.id
        ).join(Server, ExpansionPlan.server_id == Server.id)

        if server_name:
            query = query.filter(Server.name.like(f"%{server_name}%"))
        if status:
            query = query.filter(Verification.status == status)
        if start_time:
            query = query.filter(Verification.created_at >= start_time)
        if end_time:
            query = query.filter(Verification.created_at <= end_time)

        total = query.count()
        if export_all:
            verifications = query.order_by(Verification.created_at.desc()).all()
        else:
            verifications = query.order_by(Verification.created_at.desc()).offset(offset).limit(limit).all()

        return {
            'total': total,
            'items': verifications,
            'page': offset // limit + 1,
            'page_size': limit
        }

    def query_metrics(self, server_name: str = None, start_time: datetime = None,
                      end_time: datetime = None, limit: int = 1000) -> Dict:
        query = self.db.query(ResourceMetric).join(Server, ResourceMetric.server_id == Server.id)

        if server_name:
            query = query.filter(Server.name.like(f"%{server_name}%"))
        if start_time:
            query = query.filter(ResourceMetric.timestamp >= start_time)
        if end_time:
            query = query.filter(ResourceMetric.timestamp <= end_time)

        total = query.count()
        metrics = query.order_by(ResourceMetric.timestamp.desc()).limit(limit).all()

        return {
            'total': total,
            'items': metrics
        }

    def query_audit_logs(self, module: str = None, action: str = None,
                         operator: str = None, start_time: datetime = None,
                         end_time: datetime = None, limit: int = 100, offset: int = 0) -> Dict:
        query = self.db.query(AuditLog)

        if module:
            query = query.filter(AuditLog.module == module)
        if action:
            query = query.filter(AuditLog.action == action)
        if operator:
            query = query.filter(AuditLog.operator.like(f"%{operator}%"))
        if start_time:
            query = query.filter(AuditLog.timestamp >= start_time)
        if end_time:
            query = query.filter(AuditLog.timestamp <= end_time)

        total = query.count()
        logs = query.order_by(AuditLog.timestamp.desc()).offset(offset).limit(limit).all()

        return {
            'total': total,
            'items': logs,
            'page': offset // limit + 1,
            'page_size': limit
        }

    def get_server_list(self, server_type: str = None, is_active: bool = None) -> List[Server]:
        query = self.db.query(Server)
        if server_type:
            query = query.filter(Server.type == server_type)
        if is_active is not None:
            query = query.filter(Server.is_active == is_active)
        return query.order_by(Server.name).all()


class ExportService:
    def __init__(self, db: Session = None):
        self.db = db or next(get_db())
        self.export_dir = os.path.join(BASE_DIR, 'exports')
        os.makedirs(self.export_dir, exist_ok=True)

    def _get_level_names(self):
        return {
            'info': '信息',
            'warning': '警告',
            'critical': '严重',
            'fatal': '致命'
        }

    def _get_status_names(self):
        return {
            'pending': '待处理',
            'acknowledged': '已确认',
            'resolved': '已解决',
            'escalated': '已升级'
        }

    def _get_resource_names(self):
        return {
            'cpu': 'CPU',
            'memory': '内存',
            'disk': '磁盘',
            'network': '网络'
        }

    def _get_expansion_status_names(self):
        return {
            'pending': '待审批',
            'approved': '已通过',
            'rejected': '已拒绝'
        }

    def _get_order_status_names(self):
        return {
            'draft': '草稿',
            'issued': '已下达',
            'in_progress': '进行中',
            'delivered': '已交付',
            'completed': '已完成',
            'cancelled': '已取消'
        }

    def _get_verification_status_names(self):
        return {
            'pending': '待验证',
            'passed': '验证通过',
            'failed': '验证失败',
            'rolled_back': '已回滚'
        }

    def _style_header(self, ws, headers, fill_color="4472C4"):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        header_font_white = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        center_align = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        ws.row_dimensions[1].height = 25
        return thin_border

    def _style_data(self, ws, thin_border, headers_count):
        for col in range(1, headers_count + 1):
            col_letter = chr(64 + col) if col <= 26 else 'A' + chr(64 + col - 26)
            ws.column_dimensions[col_letter].width = 18

    def _write_summary_sheet(self, wb, filter_conditions: Dict, total_counts: Dict):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        ws = wb.create_sheet("筛选条件摘要")

        title_font = Font(bold=True, size=14)
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        left_align = Alignment(horizontal='left', vertical='center')

        ws.cell(row=1, column=1, value="数据导出 - 筛选条件摘要").font = title_font
        ws.merge_cells('A1:B1')
        ws.cell(row=1, column=1).alignment = Alignment(horizontal='left', vertical='center')

        ws.cell(row=3, column=1, value="导出时间:").font = header_font
        ws.cell(row=3, column=2, value=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))

        row = 5
        ws.cell(row=row, column=1, value="筛选条件").font = header_font
        ws.cell(row=row, column=1).fill = header_fill
        ws.cell(row=row, column=2, value="值").font = header_font
        ws.cell(row=row, column=2).fill = header_fill
        ws.cell(row=row, column=1).border = thin_border
        ws.cell(row=row, column=2).border = thin_border

        row = 6
        condition_items = [
            ("服务器名称", filter_conditions.get('server_name', '全部')),
            ("服务器类型", filter_conditions.get('server_type', '全部')),
            ("开始时间", filter_conditions.get('start_time', '不限').strftime('%Y-%m-%d') if isinstance(filter_conditions.get('start_time'), datetime) else filter_conditions.get('start_time', '不限')),
            ("结束时间", filter_conditions.get('end_time', '不限').strftime('%Y-%m-%d %H:%M:%S') if isinstance(filter_conditions.get('end_time'), datetime) else filter_conditions.get('end_time', '不限')),
            ("预警级别", filter_conditions.get('level', '不限')),
            ("状态筛选", filter_conditions.get('status', '不限')),
        ]

        for label, value in condition_items:
            ws.cell(row=row, column=1, value=label).border = thin_border
            ws.cell(row=row, column=2, value=str(value)).border = thin_border
            ws.cell(row=row, column=1).alignment = left_align
            ws.cell(row=row, column=2).alignment = left_align
            row += 1

        row += 1
        ws.cell(row=row, column=1, value="数据统计").font = header_font
        ws.cell(row=row, column=1).fill = header_fill
        ws.cell(row=row, column=2, value="记录数").font = header_font
        ws.cell(row=row, column=2).fill = header_fill
        ws.cell(row=row, column=1).border = thin_border
        ws.cell(row=row, column=2).border = thin_border
        row += 1

        for label, count in total_counts.items():
            ws.cell(row=row, column=1, value=label).border = thin_border
            ws.cell(row=row, column=2, value=f"{count} 条").border = thin_border
            ws.cell(row=row, column=1).alignment = left_align
            ws.cell(row=row, column=2).alignment = left_align
            row += 1

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40

    def export_alerts_excel(self, server_name: str = None, start_time: datetime = None,
                            end_time: datetime = None, level: str = None, status: str = None,
                            filter_conditions: Dict = None) -> str:
        try:
            import openpyxl

            query_service = QueryService(self.db)
            result = query_service.query_alerts(
                server_name=server_name,
                start_time=start_time,
                end_time=end_time,
                level=level,
                status=status,
                export_all=True
            )

            if filter_conditions is None:
                filter_conditions = {
                    'server_name': server_name or '全部',
                    'start_time': start_time,
                    'end_time': end_time,
                    'level': level or '不限',
                    'status': status or '不限'
                }

            filename = f"alerts_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = os.path.join(self.export_dir, filename)

            wb = openpyxl.Workbook()

            ws = wb.active
            ws.title = "预警记录"

            headers = ['预警ID', '服务器名称', '资源类型', '预警级别', '标题', '当前值', '阈值', '状态', '创建时间', '确认人', '确认时间', '解决时间']
            thin_border = self._style_header(ws, headers, "4472C4")

            level_names = self._get_level_names()
            status_names = self._get_status_names()
            resource_names = self._get_resource_names()

            items = result['items']
            if len(items) != result['total']:
                print(f"[!] 警告: 实际导出 {len(items)} 条，查询总数 {result['total']} 条")

            for row_idx, alert in enumerate(items, 2):
                server = self.db.query(Server).filter(Server.id == alert.server_id).first()
                data = [
                    alert.id,
                    server.name if server else '未知',
                    resource_names.get(alert.resource_type, alert.resource_type),
                    level_names.get(alert.alert_level, alert.alert_level),
                    alert.title,
                    f"{alert.current_value}%",
                    f"{alert.threshold_value}%",
                    status_names.get(alert.status, alert.status),
                    alert.created_at.strftime('%Y-%m-%d %H:%M:%S') if alert.created_at else '',
                    alert.acknowledged_by or '',
                    alert.acknowledged_at.strftime('%Y-%m-%d %H:%M:%S') if alert.acknowledged_at else '',
                    alert.resolved_at.strftime('%Y-%m-%d %H:%M:%S') if alert.resolved_at else ''
                ]
                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row_idx, column=col, value=value)
                    cell.border = thin_border

            self._style_data(ws, thin_border, len(headers))

            self._write_summary_sheet(wb, filter_conditions, {'预警记录': result['total']})

            wb.save(filepath)

            audit_logger.log_audit(
                self.db,
                module="export",
                action="export_alerts",
                resource_type="alert",
                operator="system",
                details=f"已导出 {len(items)}/{result['total']} 条预警记录到 {filename}"
            )

            return filepath

        except Exception as e:
            print(f"导出预警记录失败: {e}")
            raise

    def export_expansions_excel(self, server_name: str = None, start_time: datetime = None,
                                end_time: datetime = None, status: str = None,
                                filter_conditions: Dict = None) -> str:
        try:
            import openpyxl

            query_service = QueryService(self.db)
            result = query_service.query_expansions(
                server_name=server_name,
                start_time=start_time,
                end_time=end_time,
                status=status,
                export_all=True
            )

            if filter_conditions is None:
                filter_conditions = {
                    'server_name': server_name or '全部',
                    'start_time': start_time,
                    'end_time': end_time,
                    'status': status or '不限'
                }

            filename = f"expansions_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = os.path.join(self.export_dir, filename)

            wb = openpyxl.Workbook()

            ws = wb.active
            ws.title = "扩容记录"
            headers = ['方案ID', '方案名称', '服务器', '资源类型', '当前配置', '推荐配置', '数量', '预估费用', '状态', '创建人', '创建时间']
            thin_border = self._style_header(ws, headers, "70AD47")

            status_names = self._get_expansion_status_names()
            resource_names = self._get_resource_names()

            items = result['items']
            if len(items) != result['total']:
                print(f"[!] 警告: 实际导出 {len(items)} 条，查询总数 {result['total']} 条")

            for row_idx, plan in enumerate(items, 2):
                server = self.db.query(Server).filter(Server.id == plan.server_id).first()
                data = [
                    plan.id,
                    plan.plan_title,
                    server.name if server else '未知',
                    resource_names.get(plan.resource_type, plan.resource_type),
                    plan.current_spec or '',
                    plan.recommended_spec or '',
                    plan.quantity,
                    f"{plan.estimated_cost:,.2f} {plan.cost_currency}",
                    status_names.get(plan.status, plan.status),
                    plan.created_by or '',
                    plan.created_at.strftime('%Y-%m-%d %H:%M:%S') if plan.created_at else ''
                ]
                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row_idx, column=col, value=value)
                    cell.border = thin_border

            self._style_data(ws, thin_border, len(headers))

            self._write_summary_sheet(wb, filter_conditions, {'扩容方案': result['total']})

            wb.save(filepath)

            audit_logger.log_audit(
                self.db,
                module="export",
                action="export_expansions",
                resource_type="expansion_plan",
                operator="system",
                details=f"已导出 {len(items)}/{result['total']} 条扩容记录到 {filename}"
            )

            return filepath

        except Exception as e:
            print(f"导出扩容记录失败: {e}")
            raise

    def export_orders_excel(self, server_name: str = None, start_time: datetime = None,
                            end_time: datetime = None, status: str = None, supplier: str = None,
                            filter_conditions: Dict = None) -> str:
        try:
            import openpyxl

            query_service = QueryService(self.db)
            result = query_service.query_orders(
                server_name=server_name,
                start_time=start_time,
                end_time=end_time,
                status=status,
                supplier=supplier,
                export_all=True
            )

            if filter_conditions is None:
                filter_conditions = {
                    'server_name': server_name or '全部',
                    'start_time': start_time,
                    'end_time': end_time,
                    'status': status or '不限',
                    'supplier': supplier or '不限'
                }

            filename = f"orders_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = os.path.join(self.export_dir, filename)

            wb = openpyxl.Workbook()

            ws = wb.active
            ws.title = "采购订单"
            headers = ['订单号', '关联方案', '服务器', '供应商', '总金额', '币种', '状态', '交付截止', '下达人', '下达时间', '交付时间', '完成时间', '备注']
            thin_border = self._style_header(ws, headers, "ED7D31")

            status_names = self._get_order_status_names()

            items = result['items']
            if len(items) != result['total']:
                print(f"[!] 警告: 实际导出 {len(items)} 条，查询总数 {result['total']} 条")

            for row_idx, order in enumerate(items, 2):
                plan = self.db.query(ExpansionPlan).filter(ExpansionPlan.id == order.expansion_plan_id).first()
                server = self.db.query(Server).filter(Server.id == plan.server_id).first() if plan else None
                data = [
                    order.order_no,
                    plan.plan_title if plan else '未知',
                    server.name if server else '未知',
                    order.supplier,
                    f"{order.total_amount:,.2f}",
                    order.currency,
                    status_names.get(order.status, order.status),
                    order.delivery_deadline.strftime('%Y-%m-%d') if order.delivery_deadline else '',
                    order.issued_by or '',
                    order.issued_at.strftime('%Y-%m-%d %H:%M:%S') if order.issued_at else '',
                    order.delivered_at.strftime('%Y-%m-%d %H:%M:%S') if order.delivered_at else '',
                    order.completed_at.strftime('%Y-%m-%d %H:%M:%S') if order.completed_at else '',
                    order.remarks or ''
                ]
                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row_idx, column=col, value=value)
                    cell.border = thin_border

            self._style_data(ws, thin_border, len(headers))

            self._write_summary_sheet(wb, filter_conditions, {'采购订单': result['total']})

            wb.save(filepath)

            audit_logger.log_audit(
                self.db,
                module="export",
                action="export_orders",
                resource_type="purchase_order",
                operator="system",
                details=f"已导出 {len(items)}/{result['total']} 条采购订单到 {filename}"
            )

            return filepath

        except Exception as e:
            print(f"导出采购订单失败: {e}")
            raise

    def export_verifications_excel(self, server_name: str = None, start_time: datetime = None,
                                   end_time: datetime = None, status: str = None,
                                   filter_conditions: Dict = None) -> str:
        try:
            import openpyxl

            query_service = QueryService(self.db)
            result = query_service.query_verifications(
                server_name=server_name,
                start_time=start_time,
                end_time=end_time,
                status=status,
                export_all=True
            )

            if filter_conditions is None:
                filter_conditions = {
                    'server_name': server_name or '全部',
                    'start_time': start_time,
                    'end_time': end_time,
                    'status': status or '不限'
                }

            filename = f"verifications_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = os.path.join(self.export_dir, filename)

            wb = openpyxl.Workbook()

            ws = wb.active
            ws.title = "验证记录"
            headers = ['验证ID', '关联方案', '服务器', '状态', '是否回滚', '回滚原因', 'CPU前', 'CPU后', '内存前', '内存后', '磁盘前', '磁盘后', '网络前', '网络后', '验证人', '验证时间']
            thin_border = self._style_header(ws, headers, "7030A0")

            status_names = self._get_verification_status_names()

            items = result['items']
            if len(items) != result['total']:
                print(f"[!] 警告: 实际导出 {len(items)} 条，查询总数 {result['total']} 条")

            for row_idx, ver in enumerate(items, 2):
                plan = self.db.query(ExpansionPlan).filter(ExpansionPlan.id == ver.expansion_plan_id).first()
                server = self.db.query(Server).filter(Server.id == plan.server_id).first() if plan else None
                data = [
                    ver.id,
                    plan.plan_title if plan else '未知',
                    server.name if server else '未知',
                    status_names.get(ver.status, ver.status),
                    '是' if ver.is_rolled_back else '否',
                    ver.rollback_reason or '',
                    f"{ver.cpu_before}%" if ver.cpu_before else '',
                    f"{ver.cpu_after}%" if ver.cpu_after else '',
                    f"{ver.memory_before}%" if ver.memory_before else '',
                    f"{ver.memory_after}%" if ver.memory_after else '',
                    f"{ver.disk_before}%" if ver.disk_before else '',
                    f"{ver.disk_after}%" if ver.disk_after else '',
                    f"{ver.network_before}%" if ver.network_before else '',
                    f"{ver.network_after}%" if ver.network_after else '',
                    ver.verified_by or '',
                    ver.verified_at.strftime('%Y-%m-%d %H:%M:%S') if ver.verified_at else ''
                ]
                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row_idx, column=col, value=value)
                    cell.border = thin_border

            self._style_data(ws, thin_border, len(headers))

            self._write_summary_sheet(wb, filter_conditions, {'验证记录': result['total']})

            wb.save(filepath)

            audit_logger.log_audit(
                self.db,
                module="export",
                action="export_verifications",
                resource_type="verification",
                operator="system",
                details=f"已导出 {len(items)}/{result['total']} 条验证记录到 {filename}"
            )

            return filepath

        except Exception as e:
            print(f"导出验证记录失败: {e}")
            raise

    def batch_export(self, export_types: List[str], start_time: datetime = None,
                     end_time: datetime = None, server_name: str = None,
                     server_type: str = None) -> Dict:
        results = {}

        filter_conditions = {
            'server_name': server_name or '全部',
            'server_type': server_type or '全部',
            'start_time': start_time,
            'end_time': end_time
        }

        query_service = QueryService(self.db)
        total_counts = {}

        if 'alerts' in export_types:
            preview = query_service.query_alerts(server_name=server_name, start_time=start_time, end_time=end_time, limit=1)
            total_counts['预警记录'] = preview['total']
            results['alerts'] = self.export_alerts_excel(
                server_name=server_name,
                start_time=start_time,
                end_time=end_time,
                filter_conditions=filter_conditions
            )

        if 'expansions' in export_types:
            preview = query_service.query_expansions(server_name=server_name, start_time=start_time, end_time=end_time, limit=1)
            total_counts['扩容方案'] = preview['total']
            results['expansions'] = self.export_expansions_excel(
                server_name=server_name,
                start_time=start_time,
                end_time=end_time,
                filter_conditions=filter_conditions
            )

        if 'orders' in export_types:
            preview = query_service.query_orders(server_name=server_name, start_time=start_time, end_time=end_time, limit=1)
            total_counts['采购订单'] = preview['total']
            results['orders'] = self.export_orders_excel(
                server_name=server_name,
                start_time=start_time,
                end_time=end_time,
                filter_conditions=filter_conditions
            )

        if 'verifications' in export_types:
            preview = query_service.query_verifications(server_name=server_name, start_time=start_time, end_time=end_time, limit=1)
            total_counts['验证记录'] = preview['total']
            results['verifications'] = self.export_verifications_excel(
                server_name=server_name,
                start_time=start_time,
                end_time=end_time,
                filter_conditions=filter_conditions
            )

        return results

